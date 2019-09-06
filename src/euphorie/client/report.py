# coding=utf-8
"""
Report
------

The landing page for the reports, plus the logic for creating the Excel report
for the action plan.
"""

from .. import MessageFactory as _
from datetime import date
from euphorie.client import model
from euphorie.client import survey
from euphorie.client import utils
from euphorie.client.adapters.session_traversal import ITraversedSurveySession
from euphorie.client.interfaces import IClientSkinLayer
from five import grok
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from plone import api
from plone.memoize.view import memoize
from zope.i18n import translate

import logging


log = logging.getLogger(__name__)

grok.templatedir("templates")


class ReportLanding(grok.View):
    """Custom report landing page.

    This replaces the standard online view of the report with a page
    offering the RTF and XLSX download options.
    """

    grok.context(ITraversedSurveySession)
    grok.require("euphorie.client.ViewSurvey")
    grok.layer(IClientSkinLayer)
    grok.template("report_landing")
    grok.name("report_view")
    variation_class = "variation-risk-assessment"

    @property
    @memoize
    def webhelpers(self):
        return api.content.get_view("webhelpers", self.context, self.request)

    def update(self):
        if not self.webhelpers.can_view_session:
            return self.request.response.redirect(self.webhelpers.client_url)
        return super(ReportLanding, self).update()


class ActionPlanTimeline(grok.View, survey._StatusHelper):
    """Generate an excel file listing all measures.

    View name: @@timeline
    """

    grok.context(ITraversedSurveySession)
    grok.require("euphorie.client.ViewSurvey")
    grok.layer(IClientSkinLayer)
    grok.name("timeline")

    @property
    def session(self):
        return self.context.session

    def get_measures(self):
        """Find all data that should be included in the report.
           Re-use the helper methods from Status to compute the
           filtered module paths and the relevant risks.

        The data is returned as a list of tuples containing a
        :py:class:`Module <euphorie.client.model.Module>`,
        :py:class:`Risk <euphorie.client.model.Risk>` and
        :py:class:`ActionPlan <euphorie.client.model.ActionPlan>`. Each
        entry in the list will correspond to a row in the generated Excel
        file.
        """
        module_paths = self.getModulePaths()
        risk_data = self.getRisks(module_paths)
        measure_data = []
        for (module, risk) in risk_data:
            action_plan_q = self.sql_session.query(model.ActionPlan).filter(
                model.ActionPlan.risk_id == risk.id
            )
            # If the risk contains no action plan, add it as a single line
            # to the results
            action_plans = action_plan_q.all() or [None]
            for action_plan in action_plans:
                measure_data.append((module, risk, action_plan))

        # sort by 1. planning start, 2. path
        by_path_measure_data = sorted(measure_data, key=lambda x: x[1].path)
        return sorted(
            by_path_measure_data,
            key=lambda x: getattr(x[2], "planning_start", date.min) or date.min,
        )

    priority_names = {
        "low": _(u"label_timeline_priority_low", default=u"Low"),
        "medium": _(u"label_timeline_priority_medium", default=u"Medium"),
        "high": _(u"label_timeline_priority_high", default=u"High"),
    }

    columns = [
        (
            "measure",
            "planning_start",
            _("label_action_plan_start", default=u"Planning start"),
        ),
        (
            "measure",
            "planning_end",
            _("label_action_plan_end", default=u"Planning end"),
        ),
        (
            "measure",
            "action_plan",
            _(
                "label_measure_action_plan",
                default=u"General approach " u"(to eliminate or reduce the risk)",
            ),
        ),
        (
            "measure",
            "prevention_plan",
            _(
                "label_measure_prevention_plan",
                default=u"Specific action(s) required to implement " u"this approach",
            ),
        ),
        (
            "measure",
            "requirements",
            _(
                "label_measure_requirements",
                default=u"Level of expertise and/or requirements needed",
            ),
        ),
        (
            "measure",
            "responsible",
            _("label_action_plan_responsible", default=u"Who is responsible?"),
        ),
        ("measure", "budget", _("label_action_plan_budget", default=u"Budget")),
        ("module", "title", _("label_section", default=u"Section")),
        ("risk", "number", _("label_risk_number", default=u"Risk number")),
        ("risk", "title", _("report_timeline_risk_title", default=u"Risk")),
        ("risk", "priority", _("report_timeline_priority", default=u"Priority")),
        ("risk", "comment", _("report_timeline_comment", default=u"Comments")),
    ]

    def priority_name(self, priority):
        title = self.priority_names.get(priority)
        if title is not None:
            return translate(title, context=self.request)
        return priority

    def create_workbook(self):
        """Create an Excel workbook containing the all risks and measures.
        """
        t = lambda txt: translate(txt, context=self.request)
        book = Workbook()
        sheet = book.worksheets[0]
        sheet.title = t(_("report_timeline_title", default=u"Timeline"))
        survey = self.context.aq_parent

        for (column, (ntype, key, title)) in enumerate(self.columns):
            sheet.cell(row=0, column=column).value = t(title)

        row = 1
        for (module, risk, measure) in self.get_measures():
            if risk.identification in ["n/a", "yes"]:
                continue

            column = 0
            if risk.is_custom_risk:
                zodb_node = None
            else:
                zodb_node = survey.restrictedTraverse(risk.zodb_path.split("/"))
            for (ntype, key, title) in self.columns:
                value = None
                if ntype == "measure":
                    value = getattr(measure, key, None)
                elif ntype == "risk":
                    value = getattr(risk, key, None)
                    if key == "priority":
                        value = self.priority_name(value)
                    elif key == "title":
                        if (
                            getattr(zodb_node, "problem_description", None)
                            and zodb_node.problem_description.strip()
                        ):
                            value = zodb_node.problem_description
                    elif key == "number":
                        if risk.is_custom_risk:
                            num_elems = value.split(".")
                            value = u".".join([u"Ω"] + num_elems[1:])
                elif ntype == "module":
                    if key == "title":
                        if risk.is_custom_risk:
                            value = utils.get_translated_custom_risks_title(
                                self.request
                            )
                        else:
                            value = module.title
                if value is not None:
                    sheet.cell(row=row, column=column).value = value
                column += 1
            row += 1
        return book

    @property
    @memoize
    def webhelpers(self):
        return api.content.get_view("webhelpers", self.context, self.request)

    def render(self):
        if not self.webhelpers.can_view_session:
            return self.request.response.redirect(self.webhelpers.client_url)
        book = self.create_workbook()
        filename = _(
            "filename_report_timeline",
            default=u"Timeline for ${title}",
            mapping={"title": self.session.title},
        )
        filename = translate(filename, context=self.request)
        self.request.response.setHeader(
            "Content-Disposition",
            'attachment; filename="%s.xlsx"' % filename.encode("utf-8"),
        )
        self.request.response.setHeader(
            "Content-Type",
            "application/vnd.openxmlformats-" "officedocument.spreadsheetml.sheet",
        )
        return save_virtual_workbook(book)
